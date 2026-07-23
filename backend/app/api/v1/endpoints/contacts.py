import uuid
import io
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_

from app.db.session import get_db
from app.core.deps import get_current_user, require_roles
from app.models.user import User, RoleEnum
from app.models.contact import Contact, ContactStatus, Seguimiento
from app.models.audit import AuditLog
from app.schemas.contact import ContactOut, ContactCreate, ContactUpdate, ContactPage
from app.services.backup import sync_contacts_to_drive
from pydantic import BaseModel
import datetime

router = APIRouter()


class SeguimientoIn(BaseModel):
    tipo: str
    descripcion: str


class SeguimientoOut(BaseModel):
    id: uuid.UUID
    contact_id: uuid.UUID
    usuario_id: uuid.UUID
    tipo: str
    descripcion: str
    fecha: datetime.datetime

    class Config:
        from_attributes = True


@router.get("/", response_model=ContactPage)
def list_contacts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    q: Optional[str] = Query(None, description="Búsqueda por nombre, apellido, documento o teléfono"),
    estado: Optional[ContactStatus] = None,
    municipio_id: Optional[uuid.UUID] = None,
    leader_id: Optional[uuid.UUID] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=200),
):
    query = db.query(Contact)

    # Un líder solo ve sus propios contactos
    if current_user.role == RoleEnum.LIDER:
        query = query.filter(Contact.leader_id == current_user.id)

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Contact.nombre.ilike(like),
                Contact.apellido.ilike(like),
                Contact.documento.ilike(like),
                Contact.telefono.ilike(like),
            )
        )
    if estado:
        query = query.filter(Contact.estado == estado)
    if municipio_id:
        query = query.filter(Contact.municipio_id == municipio_id)
    if leader_id:
        query = query.filter(Contact.leader_id == leader_id)

    total = query.count()
    items = query.order_by(Contact.fecha_creacion.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return ContactPage(total=total, page=page, page_size=page_size, items=items)


@router.post("/", response_model=ContactOut, status_code=201)
def create_contact(
    payload: ContactCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = payload.model_dump(exclude={"tag_ids"})
    contact = Contact(**data)
    db.add(contact)
    db.flush()
    db.add(AuditLog(usuario_id=current_user.id, accion="CREATE_CONTACT", entidad="Contact", entidad_id=str(contact.id)))
    db.commit()
    db.refresh(contact)
    background_tasks.add_task(sync_contacts_to_drive)
    return contact


@router.get("/{contact_id}", response_model=ContactOut)
def get_contact(
    contact_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    return contact


@router.put("/{contact_id}", response_model=ContactOut)
def update_contact(
    contact_id: uuid.UUID,
    payload: ContactUpdate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(contact, field, value)
    db.add(AuditLog(usuario_id=current_user.id, accion="UPDATE_CONTACT", entidad="Contact", entidad_id=str(contact.id)))
    db.commit()
    db.refresh(contact)
    background_tasks.add_task(sync_contacts_to_drive)
    return contact


@router.delete("/{contact_id}", status_code=204)
def delete_contact(
    contact_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.ADMIN, RoleEnum.COORDINADOR)),
):
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    db.delete(contact)
    db.add(AuditLog(usuario_id=current_user.id, accion="DELETE_CONTACT", entidad="Contact", entidad_id=str(contact_id)))
    db.commit()
    background_tasks.add_task(sync_contacts_to_drive)


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook

    contacts = db.query(Contact).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"
    ws.append(["Nombre", "Apellido", "Documento", "Teléfono", "Correo", "Estado", "Último contacto"])
    for c in contacts:
        ws.append([
            c.nombre, c.apellido, c.documento, c.telefono, c.correo,
            c.estado.value if c.estado else "", str(c.ultimo_contacto or ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contactos.xlsx"},
    )


@router.post("/import/excel", status_code=201)
def import_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.ADMIN, RoleEnum.COORDINADOR, RoleEnum.DIGITADOR)),
):
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(file.file.read()))
    ws = wb.active
    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        nombre, apellido, documento, telefono, correo = (row + (None,) * 5)[:5]
        contact = Contact(
            nombre=nombre, apellido=apellido or "", documento=documento,
            telefono=telefono, correo=correo,
        )
        db.add(contact)
        created += 1
    db.commit()
    background_tasks.add_task(sync_contacts_to_drive)
    return {"contactos_creados": created}


# ---------- Historial de seguimiento por contacto ----------

@router.get("/{contact_id}/seguimientos", response_model=list[SeguimientoOut])
def list_seguimientos(
    contact_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return (
        db.query(Seguimiento)
        .filter(Seguimiento.contact_id == contact_id)
        .order_by(Seguimiento.fecha.desc())
        .all()
    )


@router.post("/{contact_id}/seguimientos", response_model=SeguimientoOut, status_code=201)
def create_seguimiento(
    contact_id: uuid.UUID,
    payload: SeguimientoIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    contact = db.query(Contact).filter(Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")

    seguimiento = Seguimiento(
        contact_id=contact_id, usuario_id=current_user.id,
        tipo=payload.tipo, descripcion=payload.descripcion,
    )
    contact.ultimo_contacto = datetime.datetime.utcnow()
    db.add(seguimiento)
    db.add(AuditLog(usuario_id=current_user.id, accion="ADD_SEGUIMIENTO", entidad="Contact", entidad_id=str(contact_id)))
    db.commit()
    db.refresh(seguimiento)
    return seguimiento
